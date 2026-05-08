"""
Export run2_strategy_comparison results to xlsx.

One tab per (strategy × model) combination, plus a Summary tab.
Failed docs (from run_failures) appear with blank score and error type in status column.

Usage:
    uv run --with openpyxl --with pandas evals/publish_results.py
    uv run --with openpyxl --with pandas evals/publish_results.py --out my_results.xlsx
    uv run --with openpyxl --with pandas evals/publish_results.py --project run2_strategy_comparison
"""
from __future__ import annotations

import argparse
import asyncio
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

PROJECT = "run2_strategy_comparison"
OUT_PATH = Path(__file__).parent / "results_run2.xlsx"

SHORT_STRATEGY = {
    "RecursiveSummarizer":        "Recursive",
    "RollingRefineSummarizer":    "RollingRefine",
    "MapDedupeReduceSummarizer":  "MapDedupe",
    "HierarchicalTreeSummarizer": "HierTree",
}

# Ordered for consistent tab order
RUN_ORDER = [
    ("RecursiveSummarizer",        "deepwater"),
    ("RollingRefineSummarizer",    "deepwater"),
    ("MapDedupeReduceSummarizer",  "deepwater"),
    ("HierarchicalTreeSummarizer", "deepwater"),
    ("RecursiveSummarizer",        "bywater"),
    ("RollingRefineSummarizer",    "bywater"),
    ("MapDedupeReduceSummarizer",  "bywater"),
    ("HierarchicalTreeSummarizer", "bywater"),
]

COLUMNS = [
    "source_id",
    "category",
    "token_count",
    "score",
    "status",
    "duration_s",
    "output_chars",
    "input_preview",
    "summary",
    "reference",
]

COL_WIDTHS = {
    "source_id":    22,
    "category":     16,
    "token_count":  13,
    "score":         8,
    "status":       18,
    "duration_s":   12,
    "output_chars": 13,
    "input_preview": 55,
    "summary":       70,
    "reference":     70,
}


async def fetch_data(project: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    from persist import _get_pool
    pool = await _get_pool()
    async with pool.acquire() as conn:
        # All completed runs joined with scores and document metadata
        rows = await conn.fetch("""
            SELECT
                rr.strategy,
                rr.config_id,
                rr.source_id,
                rr.output           AS summary,
                rr.output_metadata,
                er.score,
                d.text              AS source_text,
                d.reference,
                d.metadata          AS doc_meta,
                c.config            AS config_json
            FROM run_results rr
            LEFT JOIN eval_results er ON (
                er.project       = rr.project
                AND er.strategy  = rr.strategy
                AND er.config_id = rr.config_id
                AND er.source_id = rr.source_id
                AND er.eval_function = 'gemini_judge'
            )
            LEFT JOIN documents d ON (
                d.project    = rr.project
                AND d.source_id = rr.source_id
            )
            LEFT JOIN configs c ON c.config_id = rr.config_id
            WHERE rr.project = $1
            ORDER BY rr.strategy, rr.config_id, rr.source_id
        """, project)

        failure_rows = await conn.fetch("""
            SELECT strategy, config_id, source_id, error_type, token_count
            FROM run_failures
            WHERE project = $1
        """, project)

        # config_id → server alias
        config_rows = await conn.fetch("""
            SELECT DISTINCT config_id, config->>'host_alias' AS host_alias
            FROM configs
        """)

    config_to_server = {r["config_id"]: r["host_alias"] or "unknown" for r in config_rows}

    run_records = []
    for r in rows:
        meta = r["doc_meta"] or {}
        if isinstance(meta, str):
            meta = json.loads(meta)
        cfg = r["config_json"] or {}
        if isinstance(cfg, str):
            cfg = json.loads(cfg)
        om = r["output_metadata"] or {}
        if isinstance(om, str):
            om = json.loads(om)

        trace = om.get("trace", [])
        duration = round(trace[-1]["duration"], 1) if trace else None
        source_text = r["source_text"] or ""
        server = config_to_server.get(r["config_id"], "unknown")

        run_records.append({
            "strategy":      r["strategy"],
            "config_id":     r["config_id"],
            "server":        server,
            "model":         cfg.get("model", ""),
            "source_id":     r["source_id"],
            "category":      meta.get("category", ""),
            "token_count":   meta.get("token_count"),
            "score":         round(r["score"], 3) if r["score"] is not None else None,
            "status":        "ok" if r["score"] is not None else "unscored",
            "duration_s":    duration,
            "output_chars":  len(r["summary"] or ""),
            "input_preview": source_text[:400].replace("\n", " "),
            "summary":       r["summary"] or "",
            "reference":     r["reference"] or "",
        })

    failure_records = []
    for f in failure_rows:
        failure_records.append({
            "strategy":   f["strategy"],
            "config_id":  f["config_id"],
            "source_id":  f["source_id"],
            "error_type": f["error_type"],
            "token_count": f["token_count"],
        })

    return pd.DataFrame(run_records), pd.DataFrame(failure_records)


def _tab_name(strategy: str, server: str) -> str:
    short = SHORT_STRATEGY.get(strategy, strategy[:10])
    suffix = "qwen" if server == "deepwater" else "gpt"
    return f"{short}-{suffix}"


def _write_sheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    # Headers
    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = top

    # Rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, col in enumerate(COLUMNS, 1):
            val = getattr(row, col, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col in ("input_preview", "summary", "reference"):
                cell.alignment = wrap
            else:
                cell.alignment = top

    # Column widths
    for col_idx, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col]

    # Row heights for text columns
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 80

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Score color scale (column D = index 4)
    score_col = COLUMNS.index("score") + 1
    score_letter = get_column_letter(score_col)
    if len(df) > 0:
        ws.conditional_formatting.add(
            f"{score_letter}2:{score_letter}{len(df) + 1}",
            ColorScaleRule(
                start_type="num", start_value=0.0, start_color="F8696B",
                mid_type="num",   mid_value=0.5,   mid_color="FFEB84",
                end_type="num",   end_value=1.0,   end_color="63BE7B",
            ),
        )


def _write_summary(ws, runs_df: pd.DataFrame, failures_df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    top = Alignment(vertical="top")

    headers = ["strategy", "server", "model", "n_ok", "n_failed",
               "mean_score", "median_score", "std_score",
               "mean_duration_s", "p90_duration_s"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = top

    failure_counts = {}
    if not failures_df.empty:
        fc = failures_df.groupby(["strategy", "config_id"]).size()
        failure_counts = fc.to_dict()

    row_idx = 2
    for strategy, server in RUN_ORDER:
        sub = runs_df[(runs_df["strategy"] == strategy) & (runs_df["server"] == server)]
        if sub.empty:
            continue
        scored = sub[sub["score"].notna()]
        cid = sub["config_id"].iloc[0] if not sub.empty else ""
        n_failed = failure_counts.get((strategy, cid), 0)

        vals = [
            strategy,
            server,
            sub["model"].iloc[0] if not sub.empty else "",
            len(scored),
            n_failed,
            round(scored["score"].mean(), 3) if not scored.empty else None,
            round(scored["score"].median(), 3) if not scored.empty else None,
            round(scored["score"].std(), 3) if not scored.empty else None,
            round(sub["duration_s"].mean(), 1) if sub["duration_s"].notna().any() else None,
            round(sub["duration_s"].quantile(0.9), 1) if sub["duration_s"].notna().any() else None,
        ]
        for col_idx, v in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col_idx, value=v).alignment = top

        row_idx += 1

    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(h) + 4, 16)

    ws.freeze_panes = "A2"

    # Score color scale on mean_score column
    mean_col = get_column_letter(headers.index("mean_score") + 1)
    ws.conditional_formatting.add(
        f"{mean_col}2:{mean_col}{row_idx}",
        ColorScaleRule(
            start_type="num", start_value=0.0, start_color="F8696B",
            mid_type="num",   mid_value=0.5,   mid_color="FFEB84",
            end_type="num",   end_value=1.0,   end_color="63BE7B",
        ),
    )


async def main(project: str, out_path: Path) -> None:
    print(f"Fetching results for project '{project}' ...")
    runs_df, failures_df = await fetch_data(project)
    print(f"  {len(runs_df)} run results, {len(failures_df)} failure records")

    # Merge failure status into runs for docs that failed and have no run_result
    # (already-failed docs with a run_result keep status='ok'/'unscored')
    if not failures_df.empty:
        # Build a lookup: (strategy, config_id, source_id) → latest error_type
        fail_idx = failures_df.groupby(
            ["strategy", "config_id", "source_id"]
        )["error_type"].last()
        for i, row in runs_df.iterrows():
            key = (row["strategy"], row["config_id"], row["source_id"])
            if key in fail_idx and row["status"] == "unscored":
                runs_df.at[i, "status"] = fail_idx[key]

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Summary tab first
    ws_summary = wb.create_sheet("Summary")
    _write_summary(ws_summary, runs_df, failures_df)

    # One tab per run
    config_to_server = runs_df.set_index("config_id")["server"].to_dict() if not runs_df.empty else {}

    for strategy, server in RUN_ORDER:
        sub = runs_df[
            (runs_df["strategy"] == strategy) & (runs_df["server"] == server)
        ][COLUMNS].copy()

        if sub.empty:
            continue

        tab = _tab_name(strategy, server)
        ws = wb.create_sheet(tab)
        _write_sheet(ws, sub)
        print(f"  {tab}: {len(sub)} rows")

    wb.save(out_path)
    print(f"\nSaved to {out_path}")

    from persist import close_pool
    await close_pool()


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--project", default=PROJECT)
    p.add_argument("--out", type=Path, default=OUT_PATH)
    args = p.parse_args()
    asyncio.run(main(args.project, args.out))
